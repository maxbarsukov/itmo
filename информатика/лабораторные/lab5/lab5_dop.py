from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.formula.translate import Translator
from openpyxl.chart import PieChart3D, Reference
from openpyxl.formatting.rule import ColorScaleRule
import datetime

wb = Workbook()

ws = wb.active
ws.title = 'лаба 5 доп'

ws.merge_cells('B1:H1')
ws['B1'] = 'Барсуков Максим, Лабораторная работа №5, доп. задание'
ws['B1'].font = Font(color="110011", bold=True)
ws['B1'].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells('J2:N2')
ws['J2'] = f'Файл создан: {datetime.datetime.now()}'

ws['B3'] = 'Числа Фибоначчи'
ws['B3'].font = Font(italic=True)

ws['B5'] = 1
ws['B6'] = 1
ws['B7'] = '=B5 + B6'

previous = 7
NUMS = 10
for i in range(NUMS - 2):
    last = f'B{previous}'
    current = f'B{previous + 1}'
    ws[current] = Translator(ws[last].value, last).translate_formula(current)
    previous += 1

pie = PieChart3D()
data = Reference(ws, min_col=2, min_row=5, max_row=5 + NUMS)
pie.add_data(data, titles_from_data=True)
pie.set_categories(data)
pie.title = "Fibonacci"

ws.add_chart(pie, "E8")

color_scale_rule = ColorScaleRule(start_type="min",
                                  start_color="0000FF00",
                                  end_type="max",
                                  end_color="00FF0000")

ws.conditional_formatting.add(f"B5:B{5 + NUMS}", color_scale_rule)

wb.save('lab5_dop.xlsx')
